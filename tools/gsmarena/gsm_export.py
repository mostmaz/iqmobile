#!/usr/bin/env python3
"""Turn the fetched specs into a review sheet.

Sheet 1 — every listing model we found specs for, with the six fields.
Sheet 2 — what we could NOT match, so the gaps are visible rather than
          silently absent. Sorted by listing count: the top of that sheet is
          where a manual mapping buys the most.
"""
import io, json, os, re
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

HERE = os.path.dirname(os.path.abspath(__file__))
L = lambda f: json.load(io.open(os.path.join(HERE, f), encoding="utf-8"))


def short_cam(s, n=2):
    """The first n camera modules — the full string is a paragraph."""
    if not s:
        return None
    mods = re.findall(r"\d+(?:\.\d+)?\s*MP[^0-9]*(?:f/[\d.]+)?[^,]*", s)
    if not mods:
        return s[:80]
    out = []
    for m in mods[:n]:
        m = re.sub(r"\s+", " ", m).strip(" ,")
        out.append(m[:46])
    return " + ".join(out)


def charge_cell(s):
    """Wired watts, or the wireless figure clearly labelled as such.

    Apple publishes no wired wattage, so falling back silently would print
    the MagSafe pad's 15W as an iPhone's charging speed — a number a
    shopper would read as the plug in the box.
    """
    if s.get("charge_w"):
        return s["charge_w"]
    if s.get("charge_w_wireless"):
        return f'{s["charge_w_wireless"]} لاسلكي'
    return None


HEAD = ["الماركة", "الموديل", "إعلانات", "متجر iQ", "متجر الأسعار",
        "الجهاز (GSMArena)", "الشاشة (إنش)", "المعالج", "الرام (GB)",
        "البطارية (mAh)", "الشحن (W)", "الكاميرا الخلفية", "الكاميرا الأمامية",
        "دقة الشاشة", "التخزين", "أُعلن", "المصدر"]


def style(ws, widths):
    ws.freeze_panes = "A2"
    ws.sheet_view.rightToLeft = True
    fill = PatternFill("solid", fgColor="1F3350")
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    for c in ws[1]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = fill
        c.alignment = Alignment(horizontal="center", vertical="center")


if __name__ == "__main__":
    matched = L("matched.json")
    specs = {s["gsm_url"]: s for s in L("device_specs.json")}
    unmatched = L("unmatched.json")

    wb = Workbook()
    ws = wb.active
    ws.title = "المواصفات"
    ws.append(HEAD)
    n_full = 0
    for m in sorted(matched, key=lambda x: (-x["n"], x["brand"])):
        s = specs.get(m["gsm_url"])
        if not s:
            continue
        if s.get("display_inches") and s.get("battery_mah") and s.get("chipset"):
            n_full += 1
        ws.append([
            m["brand"], m["model"], m["n"], m["iq"], m["price"],
            s.get("gsm_name"), s.get("display_inches"), s.get("chipset"),
            "/".join(str(r) for r in (s.get("ram_gb") or [])) or None,
            s.get("battery_mah"), charge_cell(s),
            short_cam(s.get("camera_main")), short_cam(s.get("camera_selfie"), 1),
            s.get("display_resolution"),
            "/".join(s.get("storage_options") or []) or None,
            s.get("announced"), "https://www.gsmarena.com/" + s["gsm_url"],
        ])
    style(ws, [11, 30, 9, 8, 12, 26, 12, 30, 11, 13, 10, 40, 30, 30, 22, 20, 46])

    # The storefront gets its own sheet: it is the one shop whose product
    # pages we control end to end, so its coverage is reviewed on its own
    # rather than hunted for inside 800 marketplace rows.
    ws3 = wb.create_sheet("متجر iQ")
    ws3.append(HEAD[:2] + ["نسخ"] + HEAD[5:])
    iq_have = iq_miss = 0
    for m in sorted([x for x in matched if x["iq"] > 0], key=lambda x: (x["brand"], x["model"])):
        s = specs.get(m["gsm_url"])
        if not s:
            continue
        iq_have += 1
        ws3.append([
            m["brand"], m["model"], m["iq"],
            s.get("gsm_name"), s.get("display_inches"), s.get("chipset"),
            "/".join(str(r) for r in (s.get("ram_gb") or [])) or None,
            s.get("battery_mah"), charge_cell(s),
            short_cam(s.get("camera_main")), short_cam(s.get("camera_selfie"), 1),
            s.get("display_resolution"),
            "/".join(s.get("storage_options") or []) or None,
            s.get("announced"), "https://www.gsmarena.com/" + s["gsm_url"],
        ])
    for u in sorted([x for x in unmatched if x["iq"] > 0], key=lambda x: (x["brand"], x["model"])):
        iq_miss += 1
        ws3.append([u["brand"], u["model"], u["iq"],
                    "ملحق — لا مواصفات" if u["accessory"] else "لم نلقَ مطابقة"])
    style(ws3, [11, 30, 7, 26, 12, 30, 11, 13, 10, 40, 30, 30, 22, 20, 46])

    ws2 = wb.create_sheet("بدون مواصفات")
    ws2.append(["الماركة", "الموديل", "إعلانات", "متجر iQ", "متجر الأسعار", "ملاحظة"])
    for u in sorted(unmatched, key=lambda x: -x["n"]):
        note = ""
        low = u["model"].lower()
        if any(w in low for w in ("pencil", "case", "cover", "charger", "airpod", "buds", "شاحن", "سماعة")):
            note = "ملحق — لا مواصفات"
        elif re.search(r"[؀-ۿ]", u["model"]):
            note = "اسم غير قياسي"
        ws2.append([u["brand"], u["model"], u["n"], u["iq"], u["price"], note])
    style(ws2, [12, 34, 9, 8, 12, 22])

    out = os.path.join(HERE, "device-specs-review.xlsx")
    wb.save(out)
    print(f"wrote {out}")
    print(f"  {ws.max_row - 1} models with specs ({n_full} complete on all three core fields)")
    print(f"  {ws2.max_row - 1} models without")
    print(f"  iQ store sheet: {iq_have} with specs, {iq_miss} without")
